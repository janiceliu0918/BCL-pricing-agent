"""
Excel report generator  (companion to report_html.py).
Produces a styled .xlsx dashboard with charts and a product table.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
from openpyxl.chart  import BarChart, Reference, Series, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label  import DataLabelList
from datetime import datetime

# ── helpers ───────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border() -> Border:
    s = Side(style="thin", color="BDC3C7")
    return Border(top=s, bottom=s, left=s, right=s)

def _cell(ws, row, col, value=None, bold=False, size=11, color="000000",
          bg=None, align="left", italic=False, num_fmt=None) -> openpyxl.cell.Cell:
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold=bold, size=size, color=color, italic=italic)
    c.alignment = _align(h=align, wrap=True)
    c.border    = _border()
    if bg:
        c.fill = _fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    return c

# ── colours ───────────────────────────────────────────────────────
DARK    = "1A3A5C"
HEADER  = "16213E"
GREEN   = "27AE60"
RED     = "E74C3C"
ORANGE  = "E67E22"
BLUE    = "2980B9"
GOLD    = "F39C12"
GRAY    = "95A5A6"
LGRAY   = "F4F6F9"
ACT_BG  = "D5F5E3"
DEL_BG  = "FADBD8"
WARN_BG = "FFF3CD"
WHITE   = "FFFFFF"

CAT_COLORS = {
    "Chinese Baijiu":   "E74C3C",
    "Chinese Rice Wine":"E67E22",
    "Italian Wine":     "F39C12",
    "French Wine":      "8E44AD",
    "Spanish Wine":     "2980B9",
    "US Wine":          "1ABC9C",
    "Varietal Wine":    "27AE60",
    "Collectible Wine": "9B59B6",
    "Spirits / Other":  "95A5A6",
}


def generate(analysis: dict, output_path: str) -> str:
    wb = openpyxl.Workbook()
    _sheet_dashboard(wb, analysis)
    _sheet_products(wb, analysis)
    _sheet_recommendations(wb, analysis)
    wb.save(output_path)
    return output_path


def _sheet_dashboard(wb: openpyxl.Workbook, analysis: dict) -> None:
    ws        = wb.active
    ws.title  = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE

    for r in range(1, 80):
        ws.row_dimensions[r].height = 18
        for c in range(1, 14):
            ws.cell(r, c).fill = _fill(LGRAY)

    # col widths
    widths = {1:2, 2:16, 3:16, 4:16, 5:16, 6:16, 7:16,
              8:16, 9:16, 10:16, 11:16, 12:16, 13:2}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── title ──────────────────────────────────────────────────────
    ws.merge_cells("B1:L2")
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 34
    c = ws["B1"]
    c.value     = f"BCL PORTFOLIO PRICING DASHBOARD  ·  {datetime.now().strftime('%Y-%m-%d')}"
    c.font      = _font(bold=True, size=16, color=WHITE)
    c.fill      = _fill(DARK)
    c.alignment = _align(h="center")

    ws.merge_cells("B3:L3")
    ws.row_dimensions[3].height = 6
    ws["B3"].fill = _fill(BLUE)

    # ── KPI cards ──────────────────────────────────────────────────
    ws.row_dimensions[4].height = 10
    s = analysis["summary"]
    kpis = [
        ("TOTAL SKUs",        str(s["total_skus"]),     "Full portfolio",            BLUE,   "F0F4F8"),
        ("ACTIVE",            str(s["active"]),          f"{s['active']*100//s['total_skus']}% of portfolio", GREEN, ACT_BG),
        ("PENDING DELIST ⚠",  str(s["pending_delist"]), "All wines at risk",         RED,    DEL_BG),
        ("AVG MARKUP",        f"{s['avg_markup']:.1f}%","Range: ±%",                 GOLD,   WARN_BG),
        ("PORTFOLIO RETAIL",  f"${s['total_retail']:,.0f}", f"WHS: ${s['total_whs']:,.0f}", ORANGE, "FEF9E7"),
    ]
    for i, (label, val, sub, accent, bg) in enumerate(kpis):
        col = 2 + i * 2
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
        ws.row_dimensions[5].height = 14
        ws.row_dimensions[6].height = 38
        ws.row_dimensions[7].height = 20
        for r in range(5, 8):
            ws.cell(r, col).fill = _fill(bg)
            ws.cell(r, col).border = _border()
        lc = ws.cell(5, col, label)
        lc.font = _font(size=8, color="666666"); lc.alignment = _align(h="center")
        vc = ws.cell(6, col, val)
        vc.font = _font(bold=True, size=22, color=accent); vc.alignment = _align(h="center")
        sc = ws.cell(7, col, sub)
        sc.font = _font(size=8, color="888888", italic=True); sc.alignment = _align(h="center")

    # ── Price positioning table (wine only) ────────────────────────
    ws.row_dimensions[9].height  = 6
    ws.merge_cells("B10:L10")
    ws.row_dimensions[10].height = 22
    c = ws["B10"]
    c.value = "WINE PRICE POSITIONING  vs  BC LIQUOR MARKET"
    c.font  = _font(bold=True, size=11, color=WHITE)
    c.fill  = _fill(DARK); c.alignment = _align(h="center")

    p_hdrs = ["SKU","Description","Category","Status","WHS Final Retail",
              "BCL Min","BCL Median","BCL Max","vs Median $","vs Median %","Position"]
    hrow = 11
    ws.row_dimensions[hrow].height = 26
    for ci, h in enumerate(p_hdrs, start=2):
        c = ws.cell(hrow, ci, h)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(HEADER); c.alignment = _align(h="center"); c.border = _border()

    wine_prods = [p for p in analysis["products"]
                  if p.get("wine_type") and p.get("bcl_stats", {}).get("count", 0) > 0]
    chart_start = hrow + 1
    for ri, p in enumerate(wine_prods):
        dr = hrow + 1 + ri
        ws.row_dimensions[dr].height = 20
        bg = DEL_BG if "Delist" in p.get("status","") else ACT_BG
        row_vals = [
            p["sku"], p["description"], p["category"], p["status"],
            p.get("final_retail"), p["bcl_stats"]["min"], p["bcl_stats"]["median"],
            p["bcl_stats"]["max"], p.get("vs_median"),
            (p["vs_pct"]/100) if p.get("vs_pct") is not None else None,
            p.get("position",""),
        ]
        for ci, val in enumerate(row_vals, start=2):
            c = ws.cell(dr, ci, val)
            c.fill = _fill(bg); c.border = _border()
            c.font = _font(size=9)
            if ci in (6,7,8,9,10):
                c.number_format = '"$"#,##0.00'; c.alignment = _align(h="right")
            elif ci == 11:
                c.number_format = '+0.0%;-0.0%'; c.alignment = _align(h="center")
            elif ci == 12:
                col = GREEN if "✅" in str(val) else (RED if "🔴" in str(val) else ORANGE)
                c.font = _font(size=9, bold=True, color=col); c.alignment = _align(h="center")
            else:
                c.alignment = _align()

    chart_end = hrow + len(wine_prods)

    # ── Bar chart ──────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "bar"; chart.grouping = "clustered"
    chart.title = "Price Positioning vs BCL Market"; chart.style = 10
    chart.y_axis.title = "CAD $"; chart.width = 22; chart.height = 12

    cats = Reference(ws, min_col=3, min_row=chart_start, max_row=chart_end)
    for col, title, color in [
        (7, "BCL Min",    "AED6F1"),
        (8, "BCL Median", "2980B9"),
        (6, "Your Price", "E67E22"),
        (9, "BCL Max",    "D5D8DC"),
    ]:
        s = Series(Reference(ws, min_col=col, min_row=chart_start, max_row=chart_end), title=title)
        s.graphicalProperties.solidFill = color
        chart.series.append(s)
    chart.set_categories(cats)
    ws.add_chart(chart, f"B{chart_end+2}")

    # ── Category pie chart ─────────────────────────────────────────
    cat_start_row = chart_end + 20
    ws.merge_cells(f"B{cat_start_row}:L{cat_start_row}")
    c = ws.cell(cat_start_row, 2, "PORTFOLIO VALUE BY CATEGORY")
    c.font = _font(bold=True, size=11, color=WHITE)
    c.fill = _fill(DARK); c.alignment = _align(h="center")

    cat_hdr = cat_start_row + 1
    for ci, h in enumerate(["Category","Retail Value","Share %"], start=2):
        ws.cell(cat_hdr, ci, h).font      = _font(bold=True, size=9, color=WHITE)
        ws.cell(cat_hdr, ci, h).fill      = _fill(HEADER)
        ws.cell(cat_hdr, ci, h).alignment = _align(h="center")
        ws.cell(cat_hdr, ci, h).border    = _border()

    cat_val = analysis["summary"]["cat_value"]
    total   = sum(cat_val.values())
    for ri, (cat, val) in enumerate(sorted(cat_val.items(), key=lambda x: -x[1])):
        dr = cat_hdr + 1 + ri
        ws.cell(dr, 2, cat).fill = _fill("F4F6F9"); ws.cell(dr, 2, cat).border = _border()
        ws.cell(dr, 3, val).number_format = '"$"#,##0.00'; ws.cell(dr, 3, val).border = _border()
        pct = val / total if total else 0
        ws.cell(dr, 4, pct).number_format = "0.0%"; ws.cell(dr, 4, pct).border = _border()

    pie = PieChart()
    pie.title  = "Category Value Mix"; pie.style = 10
    pie.width  = 14; pie.height = 11
    pie_cats   = Reference(ws, min_col=2, min_row=cat_hdr+1, max_row=cat_hdr+len(cat_val))
    pie_data   = Reference(ws, min_col=3, min_row=cat_hdr,   max_row=cat_hdr+len(cat_val))
    pie.add_data(pie_data, titles_from_data=True); pie.set_categories(pie_cats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True; pie.dataLabels.showCatName = True
    ser = pie.series[0]
    for i, cat in enumerate(sorted(cat_val.items(), key=lambda x: -x[1])):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = CAT_COLORS.get(cat[0], GRAY)
        ser.dPt.append(pt)
    ws.add_chart(pie, f"F{cat_hdr}")


def _sheet_products(wb: openpyxl.Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("Product Detail")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN

    widths = {1:2, 2:10, 3:38, 4:20, 5:14, 6:14, 7:11, 8:11,
              9:16, 10:12, 11:12, 12:12, 13:14, 14:2}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(1, 50):
        ws.row_dimensions[r].height = 18
        for c in range(1, 15):
            ws.cell(r, c).fill = _fill(LGRAY)

    ws.merge_cells("B1:M2")
    ws.row_dimensions[2].height = 30
    ws["B1"].value = "PRODUCT DETAIL  ·  WHS RTL vs BCL MARKET"
    ws["B1"].font  = _font(bold=True, size=13, color=WHITE)
    ws["B1"].fill  = _fill(DARK); ws["B1"].alignment = _align(h="center")

    headers = ["SKU","Description","Category","Wholesale $","Final Retail $",
               "Markup %","Margin %","Status","BCL Min","BCL Median",
               "BCL Max","vs Median $","vs Median %"]
    hrow = 3
    ws.row_dimensions[hrow].height = 26
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(hrow, ci, h)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(HEADER); c.alignment = _align(h="center"); c.border = _border()

    for ri, p in enumerate(analysis["products"]):
        dr = hrow + 1 + ri
        ws.row_dimensions[dr].height = 20
        vs_med = p.get("vs_median")
        vs_pct = (p["vs_pct"] / 100) if p.get("vs_pct") is not None else None
        bg = DEL_BG if "Delist" in p.get("status","") else (
             WARN_BG if p.get("markup_pct") and p["markup_pct"] > 35 else ACT_BG)
        vals = [p["sku"], p["description"], p["category"], p.get("wholesale"),
                p.get("final_retail"), (p["markup_pct"]/100) if p.get("markup_pct") else None,
                (p["margin_pct"]/100) if p.get("margin_pct") else None, p.get("status",""),
                p["bcl_stats"]["min"], p["bcl_stats"]["median"], p["bcl_stats"]["max"],
                vs_med, vs_pct]
        for ci, val in enumerate(vals, start=2):
            c = ws.cell(dr, ci, val)
            c.fill = _fill(bg); c.border = _border(); c.font = _font(size=9)
            if ci in (5, 6, 10, 11, 12, 13):
                c.number_format = '"$"#,##0.00'; c.alignment = _align(h="right")
            elif ci in (7, 8):
                c.number_format = "0.0%"; c.alignment = _align(h="center")
                if val and val > 0.35:
                    c.font = _font(size=9, bold=True, color=RED)
            elif ci == 14:
                c.number_format = "+0.0%;-0.0%"; c.alignment = _align(h="center")
                if val and val < 0:
                    c.font = _font(size=9, bold=True, color=GREEN)
                elif val and val > 0:
                    c.font = _font(size=9, bold=True, color=RED)
            elif ci == 9:
                col = RED if "Delist" in str(val) else GREEN
                c.font = _font(size=9, bold=True, color=col); c.alignment = _align(h="center")
            else:
                c.alignment = _align()

    ws.auto_filter.ref = f"B{hrow}:N{hrow+len(analysis['products'])}"
    ws.freeze_panes   = f"B{hrow+1}"


def _sheet_recommendations(wb: openpyxl.Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("Recommendations")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ORANGE

    widths = {1:2, 2:18, 3:52, 4:42, 5:20, 6:2}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    for r in range(1, 40):
        ws.row_dimensions[r].height = 18
        for c in range(1, 7):
            ws.cell(r, c).fill = _fill(LGRAY)

    ws.merge_cells("B1:E2")
    ws.row_dimensions[2].height = 30
    ws["B1"].value = "ACTIONABLE RECOMMENDATIONS"
    ws["B1"].font  = _font(bold=True, size=13, color=WHITE)
    ws["B1"].fill  = _fill(DARK); ws["B1"].alignment = _align(h="center")

    headers = ["Priority","Issue","Recommendation","Owner"]
    hrow = 3
    ws.row_dimensions[hrow].height = 24
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(hrow, ci, h)
        c.font = _font(bold=True, size=9, color=WHITE)
        c.fill = _fill(HEADER); c.alignment = _align(h="center"); c.border = _border()

    s = analysis["summary"]
    products = analysis["products"]
    delist   = analysis["by_status"].get("Pending Delist", [])
    above    = [p for p in products if p.get("vs_pct") and p["vs_pct"] > 20]
    below    = [p for p in products if p.get("vs_pct") and p["vs_pct"] < -10]

    recs = []
    if delist:
        recs.append(("🔴 URGENT", DEL_BG,
            f"{len(delist)} SKUs Pending Delist ({len(delist)*100//s['total_skus']}% of portfolio)",
            "Review replacement SKUs or accept category exit",
            "Sales / Supplier Mgmt"))
    for p in above:
        recs.append(("🟡 REVIEW", WARN_BG,
            f"{p['description'][:40]} — {p['vs_pct']:+.0f}% vs BCL median",
            "Verify premium pricing justification",
            "Pricing Team"))
    for p in below:
        recs.append(("🟢 MAINTAIN", ACT_BG,
            f"{p['description'][:40]} — {abs(p['vs_pct']):.0f}% BELOW BCL median",
            "Prioritise in promotions — competitive advantage",
            "Sales / Marketing"))
    for p in s.get("outlier_markup", []):
        recs.append(("🔵 MONITOR", "D6EAF8",
            f"{p['description'][:40]} — Markup {p['markup_pct']:.1f}% vs avg {s['avg_markup']:.1f}%",
            "Verify category duty differential or flag for review",
            "Finance"))
    recs.append(("🔵 MONITOR", "D6EAF8",
        "0 products visible on BCLiquorstores.com",
        "Explore BCLDB consumer channel listing path",
        "Channel Strategy"))

    for ri, (pri, bg, issue, rec, owner) in enumerate(recs):
        dr = hrow + 1 + ri
        ws.row_dimensions[dr].height = 36
        for ci, val in enumerate([pri, issue, rec, owner], start=2):
            c = ws.cell(dr, ci, val)
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align(wrap=True)
            if ci == 2:
                c.font = _font(bold=True, size=10)
            else:
                c.font = _font(size=9)
