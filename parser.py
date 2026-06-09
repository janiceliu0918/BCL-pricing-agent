"""
WHS RTL file parser.
Supports any xlsx file whose sheet contains the BCLDB wholesale price list format.
Expected columns (detected by header, order-insensitive):
  SKU | STATUS | DESCRIPTION | % ALCOHOL | CASE CONFIGURATION |
  RETAIL PRICE PER SELLING UNIT | PROMOTION AMOUNT |
  FINAL RETAIL PRICE PER SELLING UNIT | CONTAINER DEPOSIT |
  Wholesale | Markup  [| Margin]
"""

import re
import openpyxl
from pathlib import Path

# Filename patterns that identify a WHS RTL file
WHS_FILENAME_PATTERNS = [
    r"copy\s*of\s*whs\s*rtl",
    r"whs\s*rtl",
    r"retail\s*price\s*list",
    r"retailpricelist",
]

_COMPILED = [re.compile(p, re.IGNORECASE) for p in WHS_FILENAME_PATTERNS]


def is_whs_file(path: str) -> bool:
    name = Path(path).stem
    return any(pat.search(name) for pat in _COMPILED)


def _safe_float(val) -> object:
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def _safe_str(val) -> str:
    return str(val).strip() if val is not None else ""


def _safe_int(val) -> str:
    """Return SKU as a clean integer string."""
    if val is None:
        return ""
    try:
        return str(int(float(str(val))))
    except (ValueError, TypeError):
        return _safe_str(val)


# Maps canonical field names → possible header variants (lower-case)
_HEADER_MAP = {
    "sku":          ["sku"],
    "status":       ["status"],
    "description":  ["description", "desc"],
    "alcohol":      ["% alcohol", "alcohol"],
    "case_config":  ["case configuration", "case config"],
    "retail_price": ["retail price per selling unit"],
    "promo_amount": ["promotion amount", "promo amount"],
    "final_retail": ["final retail price per selling unit",
                     "final retail price per selling",
                     "final retail"],
    "container_dep":["container deposit per selling unit",
                     "container deposit per selling",
                     "container deposit"],
    "wholesale":    ["wholesale"],
    "markup":       ["markup"],
    "margin":       ["margin"],
}


def _map_headers(raw_headers: list) -> dict:
    """Return {canonical_name: col_index} from a raw header row."""
    lower = [str(h).lower().strip() if h else "" for h in raw_headers]
    result = {}
    for field, variants in _HEADER_MAP.items():
        for i, cell in enumerate(lower):
            if any(v in cell for v in variants):
                result[field] = i
                break
    return result


def parse_whs_file(filepath: str) -> list:
    """
    Parse a WHS RTL xlsx file.
    Returns a list of product dicts with canonical keys.
    Raises ValueError if the file format is not recognised.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Try each sheet; use the first one that has a recognisable header
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        col_map = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                col_map = _map_headers(list(row))
                if "sku" in col_map and "final_retail" in col_map:
                    continue          # valid header found
                else:
                    col_map = None    # not the right sheet
                    break
            if col_map is None:
                break
            vals = list(row)
            sku  = _safe_int(vals[col_map.get("sku", 0)])
            if not sku:
                continue

            product = {
                "sku":          sku,
                "status":       _safe_str(vals[col_map["status"]])        if "status"       in col_map else "",
                "description":  _safe_str(vals[col_map["description"]])   if "description"  in col_map else "",
                "alcohol":      _safe_float(vals[col_map["alcohol"]])      if "alcohol"      in col_map else None,
                "case_config":  _safe_str(vals[col_map["case_config"]])   if "case_config"  in col_map else "",
                "retail_price": _safe_float(vals[col_map["retail_price"]]) if "retail_price" in col_map else None,
                "promo_amount": _safe_float(vals[col_map["promo_amount"]]) if "promo_amount" in col_map else 0.0,
                "final_retail": _safe_float(vals[col_map["final_retail"]]) if "final_retail" in col_map else None,
                "container_dep":_safe_float(vals[col_map["container_dep"]])if "container_dep"in col_map else None,
                "wholesale":    _safe_float(vals[col_map["wholesale"]])    if "wholesale"    in col_map else None,
                "markup":       _safe_float(vals[col_map["markup"]])       if "markup"       in col_map else None,
                "margin":       _safe_float(vals[col_map.get("margin", -1)])
                                if "margin" in col_map else None,
            }
            rows.append(product)

        if rows:
            wb.close()
            return rows

    wb.close()
    raise ValueError(
        f"No WHS RTL price list sheet found in '{filepath}'. "
        "Expected columns: SKU, STATUS, DESCRIPTION, FINAL RETAIL PRICE PER SELLING UNIT, Wholesale, Markup"
    )
