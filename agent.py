#!/usr/bin/env python3
"""
BCL Price Agent
===============
Automatically analyses BCLDB wholesale price lists (WHS RTL xlsx files)
against the live BC Liquor Stores product catalogue.

Usage
-----
# One-shot — analyse a specific file
python agent.py --file "pricing/Copy of WHS RTL.xlsx"

# Watch a folder — process any WHS RTL file that appears or changes
python agent.py --watch ./pricing

# Rebuild BCL wine cache and run
python agent.py --file "..." --refresh-cache
"""

import os
import sys
import time
import argparse
import webbrowser
from pathlib import Path
from datetime import datetime

# ── local modules ─────────────────────────────────────────────────
from bcl_api  import fetch_all_wines
from parser   import parse_whs_file, is_whs_file
from analysis import enrich
import report_html

# ── optional Excel report ─────────────────────────────────────────
try:
    import report_excel
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── config ────────────────────────────────────────────────────────
CACHE_FILE    = Path(__file__).parent / ".bcl_wines_cache.json"
CACHE_MAX_AGE = 86_400   # 24 hours in seconds
OUTPUT_DIR    = Path(__file__).parent / "reports"

# ══════════════════════════════════════════════════════════════════
# CACHE helpers
# ══════════════════════════════════════════════════════════════════

def _load_cache() -> object:
    """Return cached BCL wines if still fresh, else None."""
    import json
    if not CACHE_FILE.exists():
        return None
    age = time.time() - CACHE_FILE.stat().st_mtime
    if age > CACHE_MAX_AGE:
        return None
    try:
        with open(CACHE_FILE, encoding="utf-8") as f:
            data = json.load(f)
        print(f"  ↳ Using cached BCL catalogue ({len(data):,} wines, "
              f"{int(age/3600)}h old)")
        return data
    except Exception:
        return None


def _save_cache(wines: list) -> None:
    import json
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(wines, f)


def get_bcl_wines(force_refresh: bool = False) -> list:
    """Fetch (or load from cache) the full BCL wine catalogue."""
    if not force_refresh:
        cached = _load_cache()
        if cached:
            return cached

    print("  Fetching BCL wine catalogue (this takes ~25 s) …")
    total_ref = [0]

    def _progress(page, pages, n):
        total_ref[0] = n
        bar_len = 30
        filled  = int(bar_len * page / pages)
        bar     = "█" * filled + "░" * (bar_len - filled)
        print(f"\r  [{bar}] {page}/{pages} pages · {n:,} wines", end="", flush=True)

    wines = fetch_all_wines(progress_cb=_progress)
    print(f"\r  ✓ Fetched {len(wines):,} wines from BCL catalogue" + " " * 20)
    _save_cache(wines)
    return wines


# ══════════════════════════════════════════════════════════════════
# CORE PIPELINE
# ══════════════════════════════════════════════════════════════════

def run_pipeline(whs_path: str,
                 bcl_wines: list,
                 output_dir: Path = OUTPUT_DIR,
                 open_browser: bool = True) -> dict:
    """
    Full pipeline: parse → analyse → report.
    Returns the analysis dict.
    """
    whs_path = Path(whs_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'─'*55}")
    print(f"  Processing: {whs_path.name}")
    print(f"{'─'*55}")

    # 1. Parse
    print("  [1/3] Parsing WHS RTL file …")
    products = parse_whs_file(str(whs_path))
    print(f"        → {len(products)} products found")

    # 2. Analyse
    print("  [2/3] Running analysis …")
    analysis = enrich(products, bcl_wines)
    s = analysis["summary"]
    print(f"        → Active: {s['active']}  |  Pending Delist: {s['pending_delist']}")
    print(f"        → Avg markup: {s['avg_markup']:.1f}%  |  Portfolio retail: ${s['total_retail']:,.2f}")

    # 3. Reports
    print("  [3/3] Generating reports …")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem      = whs_path.stem.replace(" ", "_")

    # HTML
    html_path = output_dir / f"{stem}_{timestamp}.html"
    report_html.generate(
        analysis,
        str(html_path),
        source_file = whs_path.name,
        supplier    = _detect_supplier(whs_path),
        bcl_count   = len(bcl_wines),
    )
    print(f"        ✓ HTML  → {html_path}")

    # Excel (optional)
    if EXCEL_AVAILABLE:
        xlsx_path = output_dir / f"{stem}_{timestamp}.xlsx"
        report_excel.generate(analysis, str(xlsx_path))
        print(f"        ✓ Excel → {xlsx_path}")

    if open_browser:
        webbrowser.open(html_path.as_uri())

    print(f"\n  ✅ Done!\n")
    return analysis


def _detect_supplier(path: Path) -> str:
    """Try to read supplier name from sheet name or filename."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        name = wb.sheetnames[0] if wb.sheetnames else ""
        wb.close()
        # e.g. "RetailPriceList-bonvinws2" → "bonvinws2"
        if "-" in name:
            return name.split("-")[-1]
        return name or path.stem
    except Exception:
        return path.stem


# ══════════════════════════════════════════════════════════════════
# FOLDER WATCHER
# ══════════════════════════════════════════════════════════════════

def watch_folder(folder: str, bcl_wines: list, output_dir: Path) -> None:
    """
    Watch *folder* for new/modified WHS RTL xlsx files and auto-process them.
    Requires the 'watchdog' package.
    """
    try:
        from watchdog.observers import Observer
        from watchdog.events    import FileSystemEventHandler
    except ImportError:
        print("ERROR: 'watchdog' package not found.  Run: pip install watchdog")
        sys.exit(1)

    processed: set[str] = set()

    class _Handler(FileSystemEventHandler):
        def _try_process(self, path: str) -> None:
            if not path.endswith(".xlsx"):
                return
            if path in processed:
                return
            if not is_whs_file(path):
                return
            # Small delay to ensure the file is fully written
            time.sleep(1.5)
            try:
                processed.add(path)
                run_pipeline(path, bcl_wines, output_dir, open_browser=True)
            except Exception as exc:
                print(f"  ⚠ Error processing {path}: {exc}")

        def on_created(self, event):
            if not event.is_directory:
                self._try_process(event.src_path)

        def on_modified(self, event):
            if not event.is_directory:
                self._try_process(event.src_path)

    # Also process any matching files already in the folder
    folder_path = Path(folder)
    for existing in folder_path.glob("*.xlsx"):
        if is_whs_file(str(existing)):
            run_pipeline(str(existing), bcl_wines, output_dir, open_browser=False)
            processed.add(str(existing))

    observer = Observer()
    observer.schedule(_Handler(), str(folder_path), recursive=False)
    observer.start()

    print(f"\n  👀 Watching: {folder_path.resolve()}")
    print(f"     Drop any 'WHS RTL' xlsx file here to auto-generate a report.")
    print(f"     Press Ctrl+C to stop.\n")

    try:
        while True:
            time.sleep(2)
    except KeyboardInterrupt:
        observer.stop()

    observer.join()


# ══════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        prog="BCL Price Agent",
        description="Auto-analyse BCLDB wholesale price lists vs live BCL catalogue.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python agent.py --file "Copy of WHS RTL.xlsx"
  python agent.py --watch ./pricing
  python agent.py --file "..." --refresh-cache --no-browser
        """,
    )
    parser.add_argument(
        "--file", "-f", metavar="PATH",
        help="Path to a single WHS RTL xlsx file to analyse.",
    )
    parser.add_argument(
        "--watch", "-w", metavar="FOLDER",
        help="Folder to watch for WHS RTL files (daemon mode).",
    )
    parser.add_argument(
        "--output", "-o", metavar="FOLDER",
        default=str(OUTPUT_DIR),
        help=f"Output folder for reports (default: {OUTPUT_DIR}).",
    )
    parser.add_argument(
        "--refresh-cache", action="store_true",
        help="Force re-fetch of the BCL wine catalogue even if cache is fresh.",
    )
    parser.add_argument(
        "--no-browser", action="store_true",
        help="Do not auto-open the report in a browser.",
    )
    args = parser.parse_args()

    if not args.file and not args.watch:
        parser.print_help()
        sys.exit(0)

    output_dir = Path(args.output)

    print("\n  ╔══════════════════════════════════════╗")
    print("  ║   BCL Price Agent  v1.0              ║")
    print("  ╚══════════════════════════════════════╝\n")

    # Load BCL catalogue
    print("  [BCL Catalogue]")
    bcl_wines = get_bcl_wines(force_refresh=args.refresh_cache)

    if args.file:
        if not Path(args.file).exists():
            print(f"ERROR: File not found: {args.file}")
            sys.exit(1)
        run_pipeline(
            args.file,
            bcl_wines,
            output_dir,
            open_browser=not args.no_browser,
        )

    elif args.watch:
        if not Path(args.watch).is_dir():
            print(f"ERROR: Not a directory: {args.watch}")
            sys.exit(1)
        watch_folder(args.watch, bcl_wines, output_dir)


if __name__ == "__main__":
    main()
